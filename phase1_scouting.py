# ============================================================
# Phase 1: 고속 정찰 스크립트 (Calamine 엔진)
# NEO GOD Ultra Framework v2.0
# ============================================================

import sys
import io
import json
import time
import os
from typing import List, Dict, Any
from pathlib import Path

# Windows 인코딩 설정 (안전한 방식)
if sys.platform == 'win32':
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

try:
    from python_calamine import CalamineWorkbook
    CALAMINE_AVAILABLE = True
except ImportError:
    CALAMINE_AVAILABLE = False
    print("[WARNING] python-calamine not available, will use openpyxl fallback")

try:
    import psutil
except ImportError:
    print("[WARNING] psutil not available, memory estimation will be limited")
    psutil = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def hyper_speed_scout(file_path: str) -> Dict[str, Any]:
    """
    Calamine 엔진을 사용한 초고속 시트 정찰
    목표: 23MB 파일 0.5초 이내
    
    Args:
        file_path: Excel 파일 경로
    
    Returns:
        Dict: {
            'file_path': str,
            'file_size_mb': float,
            'sheets': List[Dict],
            'sheet_count': int,
            'total_rows': int,
            'total_columns': int,
            'scan_time_seconds': float,
            'memory_estimate_mb': float,
            'available_memory_mb': float,
            'engine': str
        }
    """
    if not CALAMINE_AVAILABLE:
        raise ImportError("python-calamine is required for hyper_speed_scout")
    
    start_time = time.perf_counter()
    
    # Calamine으로 워크북 열기 (Rust 엔진, 초고속)
    workbook = CalamineWorkbook.from_path(file_path)
    
    sheets_info = []
    total_rows = 0
    total_columns = 0
    
    for sheet_name in workbook.sheet_names:
        sheet = workbook.get_sheet_by_name(sheet_name)
        
        # 시트 데이터를 Python 리스트로 변환 (skip_empty_area=False로 정확한 크기 파악)
        data = sheet.to_python(skip_empty_area=False)
        
        row_count = len(data) if data else 0
        col_count = len(data[0]) if data and data[0] else 0
        
        # 헤더 추출 (첫 번째 비어있지 않은 행)
        header = None
        for row in data:
            if any(cell is not None and str(cell).strip() for cell in row):
                header = [str(cell) if cell else f'col_{i}' for i, cell in enumerate(row)]
                break
        
        # 메모리 예측 (보수적 계수 2.5 적용)
        estimated_memory_mb = (row_count * col_count * 8 * 2.5) / (1024 * 1024)
        
        # 타겟 분류
        if row_count > 10000:
            target_type = 'heavy'
        elif row_count > 1000:
            target_type = 'medium'
        else:
            target_type = 'light'
        
        sheet_info = {
            'name': sheet_name,
            'rows': row_count,
            'columns': col_count,
            'header': header,
            'target_type': target_type,
            'estimated_memory_mb': round(estimated_memory_mb, 2),
            'has_data': row_count > 0 and col_count > 0
        }
        
        sheets_info.append(sheet_info)
        total_rows += row_count
        total_columns = max(total_columns, col_count)
    
    scan_time = time.perf_counter() - start_time
    
    # 총 메모리 예측
    total_memory_mb = sum(s['estimated_memory_mb'] for s in sheets_info)
    
    # 사용 가능한 메모리 확인
    available_memory_mb = 0
    if psutil:
        available_memory_mb = psutil.virtual_memory().available / (1024 * 1024)
    
    return {
        'file_path': file_path,
        'file_size_mb': round(os.path.getsize(file_path) / (1024 * 1024), 2),
        'sheets': sheets_info,
        'sheet_count': len(sheets_info),
        'total_rows': total_rows,
        'total_columns': total_columns,
        'scan_time_seconds': round(scan_time, 4),
        'memory_estimate_mb': round(total_memory_mb, 2),
        'available_memory_mb': round(available_memory_mb, 2),
        'engine': 'calamine'
    }


def _openpyxl_fallback_scout(file_path: str) -> Dict[str, Any]:
    """
    openpyxl 폴백 (기존 로직)
    
    Args:
        file_path: Excel 파일 경로
    
    Returns:
        Dict: 스캔 결과
    """
    if openpyxl is None:
        raise ImportError("openpyxl is required for fallback scout")
    
    start_time = time.perf_counter()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    
    sheets_info = []
    total_rows = 0
    total_columns = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        row_count = sheet.max_row or 0
        col_count = sheet.max_column or 0
        
        # 타겟 분류
        if row_count > 10000:
            target_type = 'heavy'
        elif row_count > 1000:
            target_type = 'medium'
        else:
            target_type = 'light'
        
        # 메모리 예측
        estimated_memory_mb = (row_count * col_count * 8 * 2.5) / (1024 * 1024)
        
        sheets_info.append({
            'name': sheet_name,
            'rows': row_count,
            'columns': col_count,
            'target_type': target_type,
            'estimated_memory_mb': round(estimated_memory_mb, 2),
            'has_data': row_count > 0 and col_count > 0
        })
        
        total_rows += row_count
        total_columns = max(total_columns, col_count)
    
    wb.close()
    scan_time = time.perf_counter() - start_time
    
    # 사용 가능한 메모리 확인
    available_memory_mb = 0
    if psutil:
        available_memory_mb = psutil.virtual_memory().available / (1024 * 1024)
    
    return {
        'file_path': file_path,
        'file_size_mb': round(os.path.getsize(file_path) / (1024 * 1024), 2),
        'sheets': sheets_info,
        'sheet_count': len(sheets_info),
        'total_rows': total_rows,
        'total_columns': total_columns,
        'scan_time_seconds': round(scan_time, 4),
        'memory_estimate_mb': round(sum(s['estimated_memory_mb'] for s in sheets_info), 2),
        'available_memory_mb': round(available_memory_mb, 2),
        'engine': 'openpyxl_fallback'
    }


def scout_with_fallback(file_path: str) -> Dict[str, Any]:
    """
    Calamine 실패 시 openpyxl로 폴백
    
    Args:
        file_path: Excel 파일 경로
    
    Returns:
        Dict: 스캔 결과
    """
    try:
        return hyper_speed_scout(file_path)
    except Exception as e:
        print(f"[WARNING] Calamine failed: {e}, falling back to openpyxl")
        return _openpyxl_fallback_scout(file_path)


def generate_scouting_report(result: Dict[str, Any], output_dir: str = './output') -> Dict[str, str]:
    """
    정찰 리포트 생성
    
    Args:
        result: scout_with_fallback 결과
        output_dir: 출력 디렉토리
    
    Returns:
        Dict: 생성된 파일 경로들
    """
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    # 타겟 시트 분류
    targets = {
        'primary': [],
        'secondary': [],
        'excluded': []
    }
    
    for sheet in result['sheets']:
        if sheet['target_type'] == 'heavy':
            targets['primary'].append(sheet['name'])
        elif sheet['target_type'] == 'medium':
            targets['secondary'].append(sheet['name'])
        else:
            targets['excluded'].append(sheet['name'])
    
    # scouting_report.json 생성
    report = {
        'scan_time': time.strftime('%Y-%m-%dT%H:%M:%S'),
        'file_path': result['file_path'],
        'file_size_mb': result['file_size_mb'],
        'total_sheets': result['sheet_count'],
        'sheets': result['sheets'],
        'targets': targets,
        'execution_time_sec': result['scan_time_seconds'],
        'engine': result['engine'],
        'memory_estimate_mb': result['memory_estimate_mb'],
        'available_memory_mb': result['available_memory_mb']
    }
    
    report_path = output_path / 'scouting_report.json'
    with open(report_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    
    # target_sheets.json 생성
    target_sheets = {
        'primary': targets['primary'],
        'secondary': targets['secondary'],
        'all_targets': targets['primary'] + targets['secondary']
    }
    
    target_path = output_path / 'target_sheets.json'
    with open(target_path, 'w', encoding='utf-8') as f:
        json.dump(target_sheets, f, ensure_ascii=False, indent=2)
    
    print(f"[완료] 정찰 리포트 생성:")
    print(f"  - {report_path}")
    print(f"  - {target_path}")
    print(f"  - 스캔 시간: {result['scan_time_seconds']}초")
    print(f"  - 엔진: {result['engine']}")
    print(f"  - 타겟 시트: {len(target_sheets['all_targets'])}개")
    
    return {
        'scouting_report': str(report_path),
        'target_sheets': str(target_path)
    }


if __name__ == '__main__':
    # 테스트 실행
    import argparse
    
    parser = argparse.ArgumentParser(description='Phase 1: 고속 정찰')
    parser.add_argument('file_path', type=str, help='Excel 파일 경로')
    parser.add_argument('--output-dir', type=str, default='./output', help='출력 디렉토리')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Phase 1: 고속 정찰 (Hyper-Speed Scouting)")
    print("=" * 60)
    print(f"파일: {args.file_path}")
    print()
    
    result = scout_with_fallback(args.file_path)
    generate_scouting_report(result, args.output_dir)
