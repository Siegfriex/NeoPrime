# ============================================================
# Phase 2: 물리적 이원화 추출 (Physical Dual-Track Extraction)
# NEO GOD Ultra Framework v2.0
# ============================================================

import sys
import io
import json
import time
import re
from typing import Dict, List, Any, Generator, Tuple, Optional
from pathlib import Path

# Windows 인코딩 설정 (안전한 방식)
if sys.platform == 'win32':
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

try:
    import openpyxl
except ImportError:
    openpyxl = None
    print("[WARNING] openpyxl not available, formula extraction will fail")


class PhysicalDualTrackExtractor:
    """
    물리적으로 분리된 이원화 추출기
    
    Track A: 값(Value) 추출 - Pandas + Calamine (초고속)
    Track B: 수식(Formula) 추출 - OpenPyXL (샘플링 10행)
    """
    
    def __init__(self, file_path: str, output_dir: str = './output'):
        """
        Args:
            file_path: Excel 파일 경로
            output_dir: 출력 디렉토리
        """
        self.file_path = file_path
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    # ========================
    # Track A: 값 추출 (고속)
    # ========================
    
    def extract_values_with_calamine(
        self, 
        sheet_name: str, 
        chunk_size: int = 50000
    ) -> Generator[Tuple[pd.DataFrame, int], None, None]:
        """
        Calamine 엔진으로 값 추출 (초고속)
        청크 단위 Generator 반환으로 메모리 효율 극대화
        
        Args:
            sheet_name: 시트 이름
            chunk_size: 청크 크기 (행 수)
        
        Yields:
            Tuple[pd.DataFrame, int]: (청크 데이터프레임, 청크 인덱스)
        """
        # Pandas 2.2+ Calamine 직접 지원
        try:
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine='calamine',  # Rust 기반 초고속 엔진
                dtype=str  # 모든 데이터를 문자열로 읽어 타입 손실 방지
            )
        except Exception as e:
            print(f"[WARNING] Calamine failed: {e}, using openpyxl")
            df = pd.read_excel(
                self.file_path,
                sheet_name=sheet_name,
                engine='openpyxl',
                dtype=str
            )
        
        # 청크 단위 분할 반환
        total_rows = len(df)
        for chunk_idx, start_row in enumerate(range(0, total_rows, chunk_size)):
            end_row = min(start_row + chunk_size, total_rows)
            chunk_df = df.iloc[start_row:end_row].copy()
            
            # 청크 메타데이터 추가
            chunk_df.attrs['chunk_index'] = chunk_idx
            chunk_df.attrs['row_range'] = (start_row, end_row)
            
            yield chunk_df, chunk_idx
    
    # ========================
    # Track B: 수식 추출 (샘플링)
    # ========================
    
    def extract_formula_samples(
        self, 
        sheet_name: str, 
        sample_rows: int = 10
    ) -> Dict[str, Any]:
        """
        수식 메타데이터 샘플 추출 (OpenPyXL)
        
        핵심 전략: 20만 행 전체 수식 추출은 비효율적
        → 헤더 + 상위 N행의 "대표 수식"만 추출하여 메타데이터화
        
        Args:
            sheet_name: 시트 이름
            sample_rows: 샘플링할 행 수 (기본 10행)
        
        Returns:
            Dict: 수식 메타데이터
        """
        if openpyxl is None:
            return {
                'sheet_name': sheet_name,
                'columns_with_formulas': [],
                'formula_samples': {},
                'formula_patterns': {},
                'error': 'openpyxl not available'
            }
        
        wb = openpyxl.load_workbook(
            self.file_path, 
            read_only=True, 
            data_only=False  # 수식 문자열 추출을 위해 False
        )
        sheet = wb[sheet_name]
        
        formula_metadata = {
            'sheet_name': sheet_name,
            'columns_with_formulas': [],
            'formula_samples': {},
            'formula_patterns': {}
        }
        
        # 헤더 추출 (첫 번째 행)
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        headers = [cell.value if cell.value else f'col_{i}' for i, cell in enumerate(header_row)]
        
        # 샘플 행에서 수식 추출
        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=2, max_row=sample_rows + 1, values_only=False), 
            start=2
        ):
            for col_idx, cell in enumerate(row):
                # 수식 셀 감지 (data_type='f' 또는 값이 '='로 시작)
                is_formula = False
                formula_str = None
                
                if cell.data_type == 'f':
                    is_formula = True
                    formula_str = str(cell.value)
                elif isinstance(cell.value, str) and cell.value.startswith('='):
                    is_formula = True
                    formula_str = cell.value
                
                if is_formula and formula_str:
                    col_name = headers[col_idx] if col_idx < len(headers) else f'col_{col_idx}'
                    
                    # 해당 컬럼의 첫 수식만 저장 (대표 수식)
                    if col_name not in formula_metadata['formula_samples']:
                        formula_metadata['formula_samples'][col_name] = {
                            'sample_row': row_idx,
                            'formula': formula_str,
                            'dependencies': self._extract_cell_references(formula_str)
                        }
                        formula_metadata['columns_with_formulas'].append(col_name)
                        
                        # 수식 패턴 분류
                        pattern = self._classify_formula_pattern(formula_str)
                        formula_metadata['formula_patterns'][col_name] = pattern
        
        wb.close()
        return formula_metadata
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """
        수식에서 셀 참조 추출 (개선된 정규식)
        
        지원 패턴:
        - 상대 참조: A1, B2
        - 절대 참조: $A$1, $B$2
        - 혼합 참조: $A1, A$1
        - 범위 참조: A1:B10
        - 시트 참조: Sheet1!A1
        
        Args:
            formula: 수식 문자열
        
        Returns:
            List[str]: 셀 참조 리스트
        """
        if not formula or not isinstance(formula, str):
            return []
        
        patterns = [
            r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)",  # Sheet!Cell
            r'\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)',  # Range A1:B10
            r'\$?([A-Z]+)\$?(\d+)',  # Single cell
        ]
        
        references = []
        for pattern in patterns:
            matches = re.findall(pattern, formula.upper())
            for match in matches:
                if isinstance(match, tuple):
                    ref = ''.join(str(m) for m in match if m)
                else:
                    ref = match
                if ref and ref not in references:
                    references.append(ref)
        
        return references
    
    def _classify_formula_pattern(self, formula: str) -> str:
        """
        수식 패턴 분류
        
        Args:
            formula: 수식 문자열
        
        Returns:
            str: 패턴 타입
        """
        formula_upper = formula.upper()
        
        if 'SUM(' in formula_upper:
            return 'aggregation_sum'
        elif 'AVERAGE(' in formula_upper or 'AVG(' in formula_upper:
            return 'aggregation_avg'
        elif 'COUNT(' in formula_upper:
            return 'aggregation_count'
        elif 'IF(' in formula_upper:
            return 'conditional'
        elif 'VLOOKUP(' in formula_upper or 'HLOOKUP(' in formula_upper:
            return 'lookup'
        elif 'INDEX(' in formula_upper or 'MATCH(' in formula_upper:
            return 'index_match'
        else:
            return 'other'
    
    # ========================
    # 통합 실행
    # ========================
    
    def run_dual_track_extraction(
        self, 
        sheet_name: str,
        chunk_size: int = 50000,
        formula_sample_rows: int = 10
    ) -> Dict[str, Any]:
        """
        이원화 추출 통합 실행
        
        Args:
            sheet_name: 시트 이름
            chunk_size: 청크 크기
            formula_sample_rows: 수식 샘플링 행 수
        
        Returns:
            Dict: {
                'value_chunks': List[str],  # 저장된 Parquet 파일 경로들
                'formula_metadata': Dict,    # 수식 메타데이터
                'stats': Dict               # 추출 통계
            }
        """
        result = {
            'value_chunks': [],
            'formula_metadata': None,
            'stats': {
                'total_rows': 0,
                'total_chunks': 0,
                'extraction_time_seconds': 0
            }
        }
        
        start_time = time.perf_counter()
        
        # Track A: 값 추출 (Calamine)
        print(f"[Track A] 값 추출 시작 (Calamine 엔진)...")
        for chunk_df, chunk_idx in self.extract_values_with_calamine(sheet_name, chunk_size):
            # Parquet으로 저장 (Snappy 압축)
            chunk_path = self.output_dir / f"{sheet_name}_chunk_{chunk_idx:04d}.parquet"
            
            table = pa.Table.from_pandas(chunk_df, preserve_index=False)
            pq.write_table(table, chunk_path, compression='snappy')
            
            result['value_chunks'].append(str(chunk_path))
            result['stats']['total_rows'] += len(chunk_df)
            result['stats']['total_chunks'] += 1
            
            print(f"  [Chunk {chunk_idx}] {len(chunk_df)} rows saved to {chunk_path}")
        
        # Track B: 수식 샘플 추출 (OpenPyXL)
        print(f"[Track B] 수식 샘플 추출 시작 (OpenPyXL)...")
        result['formula_metadata'] = self.extract_formula_samples(sheet_name, formula_sample_rows)
        
        # 수식 메타데이터 저장
        formula_path = self.output_dir / f"{sheet_name}_formula_metadata.json"
        with open(formula_path, 'w', encoding='utf-8') as f:
            json.dump(result['formula_metadata'], f, ensure_ascii=False, indent=2)
        
        result['stats']['extraction_time_seconds'] = round(time.perf_counter() - start_time, 2)
        
        print(f"[완료] 총 {result['stats']['total_rows']}행, {result['stats']['total_chunks']}청크")
        print(f"  수식 컬럼: {len(result['formula_metadata']['columns_with_formulas'])}개")
        
        return result


if __name__ == '__main__':
    # 테스트 실행
    import argparse
    
    parser = argparse.ArgumentParser(description='Phase 2: 물리적 이원화 추출')
    parser.add_argument('file_path', type=str, help='Excel 파일 경로')
    parser.add_argument('sheet_name', type=str, help='시트 이름')
    parser.add_argument('--output-dir', type=str, default='./output', help='출력 디렉토리')
    parser.add_argument('--chunk-size', type=int, default=50000, help='청크 크기')
    parser.add_argument('--formula-sample-rows', type=int, default=10, help='수식 샘플링 행 수')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Phase 2: 물리적 이원화 추출 (Physical Dual-Track Extraction)")
    print("=" * 60)
    print(f"파일: {args.file_path}")
    print(f"시트: {args.sheet_name}")
    print()
    
    extractor = PhysicalDualTrackExtractor(args.file_path, args.output_dir)
    result = extractor.run_dual_track_extraction(
        args.sheet_name,
        chunk_size=args.chunk_size,
        formula_sample_rows=args.formula_sample_rows
    )
    
    print()
    print("추출 결과:")
    print(json.dumps(result['stats'], indent=2, ensure_ascii=False))
