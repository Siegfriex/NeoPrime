"""
openpyxl을 사용한 메타데이터 추출

네임드레인지, 테이블, 데이터검증, 조건부서식, 외부링크 등의
의미 해석에 필요한 메타데이터를 openpyxl로 추출합니다.
"""

from pathlib import Path
from typing import Dict, List, Optional
from collections import defaultdict
import json
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
import zipfile
import xml.etree.ElementTree as ET

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class OpenPyXLMetadataExtractor:
    """openpyxl로 메타데이터 추출 클래스"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.metadata = {
            "named_ranges": [],
            "tables": [],
            "data_validations": [],
            "conditional_formats": [],
            "external_links": [],
        }
    
    def extract(self) -> Dict:
        """메타데이터 추출 실행"""
        logger.info(f"메타데이터 추출 시작: {self.excel_path}")
        
        self.wb = load_workbook(self.excel_path, data_only=False, read_only=True)
        
        self._extract_named_ranges()
        self._extract_tables()
        # DV/CF는 read_only openpyxl에서 누락되는 경우가 많아 XML에서 직접 추출
        self._extract_data_validations_xml()
        self._extract_conditional_formats_xml()
        self._extract_external_links()
        
        self.wb.close()
        
        logger.info("메타데이터 추출 완료")
        return self.metadata
    
    def _extract_named_ranges(self):
        """네임드레인지 추출"""
        logger.info("네임드레인지 추출 중...")
        
        for name_obj in self.wb.defined_names:
            name_data = {
                "name": name_obj.name,
                "value": str(name_obj.value) if hasattr(name_obj, 'value') else str(name_obj),
                "scope": name_obj.scope if hasattr(name_obj, 'scope') else None,
            }
            
            # 타입 추정
            value_str = name_data["value"]
            if '!' in value_str:
                name_data["type"] = "range"
            elif any(op in value_str for op in ['+', '-', '*', '/', '=']):
                name_data["type"] = "formula"
            else:
                name_data["type"] = "constant"
            
            # destinations()로 실제 범위 추출 시도
            try:
                destinations = list(name_obj.destinations)
                if destinations:
                    name_data["destinations"] = [
                        {"sheet": sheet, "range": range_ref}
                        for sheet, range_ref in destinations
                    ]
            except:
                pass
            
            self.metadata["named_ranges"].append(name_data)
        
        logger.info(f"네임드레인지 {len(self.metadata['named_ranges'])}개 추출")
    
    def _extract_tables(self):
        """테이블(ListObject) 추출"""
        logger.info("테이블 추출 중...")
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            if hasattr(ws, 'tables') and ws.tables:
                for table_name, table in ws.tables.items():
                    table_data = {
                        "sheet": sheet_name,
                        "name": table_name,
                        "ref": table.ref,
                        "display_name": table.displayName if hasattr(table, 'displayName') else table_name,
                        "columns": [],
                    }
                    
                    # 컬럼 정보
                    if hasattr(table, 'tableColumns') and table.tableColumns:
                        for col in table.tableColumns:
                            col_data = {
                                "id": col.id if hasattr(col, 'id') else None,
                                "name": col.name if hasattr(col, 'name') else None,
                            }
                            table_data["columns"].append(col_data)
                    
                    self.metadata["tables"].append(table_data)
        
        logger.info(f"테이블 {len(self.metadata['tables'])}개 추출")
    
    def _build_sheet_name_map_from_rels(self, z: zipfile.ZipFile) -> Dict[str, str]:
        """workbook.xml + workbook.xml.rels로 sheet file -> sheet name 매핑 생성"""
        sheet_name_map: Dict[str, str] = {}
        try:
            workbook_xml = z.read('xl/workbook.xml')
            rels_xml = z.read('xl/_rels/workbook.xml.rels')
            wb_root = ET.fromstring(workbook_xml)
            rels_root = ET.fromstring(rels_xml)

            ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            relns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

            rel_map: Dict[str, str] = {}
            for rel in rels_root.iter():
                if rel.tag.endswith('Relationship'):
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    if target and 'worksheets' in target:
                        if not target.startswith('xl/'):
                            target = 'xl/' + target
                        rel_map[rid] = target

            for sh in wb_root.findall('.//m:sheets/m:sheet', ns):
                name = sh.get('name')
                rid = sh.get(f'{{{relns}}}id')
                target = rel_map.get(rid)
                if name and target:
                    sheet_name_map[target] = name
        except Exception as e:
            logger.warning(f"시트명 매핑 생성 실패: {e}")
        return sheet_name_map

    def _extract_data_validations_xml(self):
        """데이터검증(XML) 추출"""
        logger.info("데이터검증(XML) 추출 중...")

        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        dvs: List[Dict] = []

        with zipfile.ZipFile(self.excel_path, 'r') as z:
            sheet_name_map = self._build_sheet_name_map_from_rels(z)
            sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
            for sheet_file in sheet_files:
                sheet_name = sheet_name_map.get(sheet_file, Path(sheet_file).stem)
                try:
                    root = ET.fromstring(z.read(sheet_file))
                except Exception:
                    continue

                for dv in root.findall('.//m:dataValidations/m:dataValidation', ns):
                    sqref = dv.get('sqref')
                    dv_data = {
                        "sheet": sheet_name,
                        "sqref": sqref,
                        "type": dv.get("type"),
                        "operator": dv.get("operator"),
                        "allowBlank": dv.get("allowBlank"),
                        "showDropDown": dv.get("showDropDown"),
                        "showInputMessage": dv.get("showInputMessage"),
                        "showErrorMessage": dv.get("showErrorMessage"),
                        "errorStyle": dv.get("errorStyle"),
                        "formula1": (dv.findtext('m:formula1', default=None, namespaces=ns)),
                        "formula2": (dv.findtext('m:formula2', default=None, namespaces=ns)),
                        "promptTitle": (dv.findtext('m:promptTitle', default=None, namespaces=ns)),
                        "prompt": (dv.findtext('m:prompt', default=None, namespaces=ns)),
                        "errorTitle": (dv.findtext('m:errorTitle', default=None, namespaces=ns)),
                        "error": (dv.findtext('m:error', default=None, namespaces=ns)),
                    }
                    dvs.append(dv_data)

        self.metadata["data_validations"] = dvs
        logger.info(f"데이터검증(XML) {len(dvs)}개 추출")
    
    def _extract_conditional_formats_xml(self):
        """조건부서식(XML) 추출"""
        logger.info("조건부서식(XML) 추출 중...")

        ns = {'m': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        cfs: List[Dict] = []

        with zipfile.ZipFile(self.excel_path, 'r') as z:
            sheet_name_map = self._build_sheet_name_map_from_rels(z)
            sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
            for sheet_file in sheet_files:
                sheet_name = sheet_name_map.get(sheet_file, Path(sheet_file).stem)
                try:
                    root = ET.fromstring(z.read(sheet_file))
                except Exception:
                    continue

                for cf in root.findall('.//m:conditionalFormatting', ns):
                    sqref = cf.get('sqref')
                    for rule in cf.findall('m:cfRule', ns):
                        formulas = [f.text for f in rule.findall('m:formula', ns) if f.text]
                        cf_data = {
                            "sheet": sheet_name,
                            "sqref": sqref,
                            "type": rule.get("type"),
                            "operator": rule.get("operator"),
                            "priority": rule.get("priority"),
                            "stopIfTrue": rule.get("stopIfTrue"),
                            "dxfId": rule.get("dxfId"),
                            "formulas": formulas,
                        }
                        cfs.append(cf_data)

        self.metadata["conditional_formats"] = cfs
        logger.info(f"조건부서식(XML) {len(cfs)}개 추출")
    
    def _extract_external_links(self):
        """외부 링크 추출"""
        logger.info("외부 링크 확인 중...")
        
        # openpyxl에서는 외부 링크 정보를 직접 제공하지 않음
        # zip 파일에서 직접 확인하는 것이 더 정확
        import zipfile
        
        with zipfile.ZipFile(self.excel_path, 'r') as z:
            external_link_files = [
                f for f in z.namelist() 
                if f.startswith('xl/externalLinks/')
            ]
            
            for link_file in external_link_files:
                link_data = {
                    "file": link_file,
                    "note": "외부 링크 파일 존재",
                }
                self.metadata["external_links"].append(link_data)
        
        logger.info(f"외부 링크 파일 {len(self.metadata['external_links'])}개 발견")
    
    def save_all(self, output_dir: str = "outputs"):
        """모든 메타데이터 저장"""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # 네임드레인지
        if self.metadata["named_ranges"]:
            with open(output_path / "named_ranges.json", 'w', encoding='utf-8') as f:
                json.dump(self.metadata["named_ranges"], f, ensure_ascii=False, indent=2)
            logger.info(f"네임드레인지 저장: {output_path / 'named_ranges.json'}")
        
        # 테이블
        if self.metadata["tables"]:
            with open(output_path / "tables.json", 'w', encoding='utf-8') as f:
                json.dump(self.metadata["tables"], f, ensure_ascii=False, indent=2)
            logger.info(f"테이블 저장: {output_path / 'tables.json'}")
        
        # 데이터검증
        if self.metadata["data_validations"]:
            with open(output_path / "data_validations.json", 'w', encoding='utf-8') as f:
                json.dump(self.metadata["data_validations"], f, ensure_ascii=False, indent=2, default=str)
            logger.info(f"데이터검증 저장: {output_path / 'data_validations.json'}")
        
        # 조건부서식
        if self.metadata["conditional_formats"]:
            with open(output_path / "conditional_formats.json", 'w', encoding='utf-8') as f:
                json.dump(self.metadata["conditional_formats"], f, ensure_ascii=False, indent=2)
            logger.info(f"조건부서식 저장: {output_path / 'conditional_formats.json'}")
        
        # 외부 링크
        if self.metadata["external_links"]:
            with open(output_path / "external_links.json", 'w', encoding='utf-8') as f:
                json.dump(self.metadata["external_links"], f, ensure_ascii=False, indent=2)
            logger.info(f"외부 링크 저장: {output_path / 'external_links.json'}")


def main():
    """메인 함수"""
    excel_path = r"C:\Neoprime\202511고속성장분석기(가채점)20251114 (1).xlsx"
    
    extractor = OpenPyXLMetadataExtractor(excel_path)
    metadata = extractor.extract()
    extractor.save_all()
    
    print("\n메타데이터 추출 완료!")
    print(f"네임드레인지: {len(metadata['named_ranges'])}개")
    print(f"테이블: {len(metadata['tables'])}개")
    print(f"데이터검증: {len(metadata['data_validations'])}개")
    print(f"조건부서식: {len(metadata['conditional_formats'])}개")
    print(f"외부 링크: {len(metadata['external_links'])}개")


if __name__ == "__main__":
    main()
