"""
XLSX XML 레벨에서 수식 고속 추출

XLSX 파일은 zip 압축 내부에 XML 파일로 저장됩니다.
xl/worksheets/sheetN.xml에 수식이 <f> 태그로 저장되므로,
모든 셀을 순회하지 않고도 수식만 빠르게 추출할 수 있습니다.
"""

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from collections import defaultdict
import pandas as pd
import logging
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class XLSXFormulaExtractor:
    """XLSX XML에서 수식 추출 클래스"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.sheet_name_map: Dict[str, str] = {}  # sheetN.xml -> 실제 시트명
        self.formulas: List[Dict] = []
        self.shared_formula_groups: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    
    def extract(self) -> pd.DataFrame:
        """수식 추출 실행"""
        logger.info(f"수식 추출 시작: {self.excel_path}")
        
        # openpyxl로 시트명 가져오기 (더 확실한 방법)
        from openpyxl import load_workbook
        wb = load_workbook(self.excel_path, read_only=True, data_only=False)
        sheet_names = wb.sheetnames
        wb.close()
        
        with zipfile.ZipFile(self.excel_path, 'r') as z:
            # 1. 시트명 매핑 생성
            self._build_sheet_name_map(z, sheet_names)
            
            # 2. 수식 추출
            sheet_files = sorted([f for f in z.namelist() 
                          if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')])
            
            logger.info(f"{len(sheet_files)}개 시트 파일 발견")
            
            for idx, sheet_file in enumerate(sheet_files):
                # 시트명 매핑이 없으면 인덱스로 매핑
                if sheet_file not in self.sheet_name_map and idx < len(sheet_names):
                    self.sheet_name_map[sheet_file] = sheet_names[idx]
                
                self._extract_formulas_from_sheet(z, sheet_file)
        
        logger.info(f"총 {len(self.formulas)}개 수식 추출 완료")
        
        # DataFrame으로 변환
        if self.formulas:
            df = pd.DataFrame(self.formulas)
        else:
            df = pd.DataFrame()
        return df
    
    def _build_sheet_name_map(self, z: zipfile.ZipFile, sheet_names: List[str]):
        """workbook.xml에서 시트 ID와 이름 매핑 생성"""
        try:
            workbook_xml = z.read('xl/workbook.xml')
            root = ET.fromstring(workbook_xml)
            
            # 네임스페이스 제거하고 직접 찾기
            for elem in root.iter():
                if '}' in elem.tag:
                    ns, tag = elem.tag.split('}', 1)
                else:
                    tag = elem.tag
            
            # relationship 파일에서 시트 파일명 매핑 찾기
            rels_file = 'xl/_rels/workbook.xml.rels'
            rel_map = {}
            if rels_file in z.namelist():
                rels_xml = z.read(rels_file)
                rels_root = ET.fromstring(rels_xml)
                
                for rel in rels_root.iter():
                    if 'Relationship' in rel.tag:
                        rel_id = rel.get('Id')
                        target = rel.get('Target')
                        if target and 'worksheets' in target:
                            rel_map[rel_id] = target
            
            # workbook.xml에서 시트 찾기
            for elem in root.iter():
                if 'sheet' in elem.tag.lower():
                    r_id = None
                    for attr_name, attr_value in elem.attrib.items():
                        if 'id' in attr_name.lower():
                            r_id = attr_value
                            break
                    
                    sheet_name = elem.get('name')
                    if not sheet_name and r_id:
                        # 인덱스로 찾기
                        try:
                            idx = int(r_id.replace('rId', '')) - 1
                            if 0 <= idx < len(sheet_names):
                                sheet_name = sheet_names[idx]
                        except:
                            pass
                    
                    if r_id in rel_map and sheet_name:
                        sheet_file = rel_map[r_id]
                        if not sheet_file.startswith('xl/'):
                            sheet_file = 'xl/' + sheet_file
                        self.sheet_name_map[sheet_file] = sheet_name
        
        except Exception as e:
            logger.warning(f"시트명 매핑 생성 실패 (인덱스로 대체): {e}")
    
    def _extract_formulas_from_sheet(self, z: zipfile.ZipFile, sheet_file: str):
        """특정 시트에서 수식 추출"""
        try:
            xml_content = z.read(sheet_file)
            root = ET.fromstring(xml_content)
            
            # 네임스페이스 정의
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            sheet_name = self.sheet_name_map.get(sheet_file, Path(sheet_file).stem)
            start_idx = len(self.formulas)
            
            # 공유수식 마스터 저장: si -> {formula, master_cell, ref}
            shared_masters: Dict[str, Dict] = {}
            sheet_shared_groups: Dict[str, List[str]] = defaultdict(list)
            
            # 수식이 있는 셀만 찾기 (더 효율적)
            formula_cells = root.findall('.//main:c[main:f]', ns)
            
            for cell in formula_cells:
                cell_ref = cell.get('r')
                if not cell_ref:
                    continue
                
                formula_elem = cell.find('.//main:f', ns)
                if formula_elem is None:
                    continue
                
                # NOTE: shared formula non-master 셀은 <f t="shared" si="..."/> 형태로 text가 비어있음
                #       이 경우에도 “수식 셀”로 카탈로그에 남겨야 그룹/확장이 가능함.
                raw_text = formula_elem.text or ""
                t_attr = formula_elem.get('t', '')  # shared/array/...
                shared_si = formula_elem.get('si')
                shared_ref = formula_elem.get('ref')  # master shared formula range

                formula_text = raw_text
                if formula_text and not formula_text.startswith('='):
                    formula_text = '=' + formula_text
                # normal/array인데 텍스트가 비어있으면 스킵 (shared는 허용)
                if not formula_text and t_attr != 'shared':
                    continue
                    
                # 행 번호 추출
                row_idx = self._cell_ref_to_row(cell_ref)
                
                formula_type = 'normal'
                if t_attr == 'shared':
                    formula_type = 'shared'
                    # 공유수식 마스터 저장: text가 있고 ref가 있으면 master로 간주
                    if shared_si is not None:
                        if shared_si not in shared_masters and formula_text:
                            shared_masters[shared_si] = {
                                'formula': formula_text,
                                'master_cell': cell_ref,
                                'ref': shared_ref,
                            }
                elif t_attr == 'array':
                    formula_type = 'array'
                
                # 외부 참조 확인
                has_external_ref = self._check_external_ref(formula_text)
                
                formula_data = {
                    'sheet_name': sheet_name,
                    'cell_ref': cell_ref,
                    'formula': formula_text,
                    'formula_type': formula_type,
                    'shared_si': shared_si,
                    'has_external_ref': has_external_ref,
                    'row': row_idx,
                    'col': self._cell_ref_to_col(cell_ref),
                    # shared meta (채워질 수 있음)
                    'shared_ref_range': shared_ref,
                    'shared_master': None,
                    'shared_group_size': None,
                }
                
                self.formulas.append(formula_data)
                
                # 공유수식 그룹에 추가
                if shared_si is not None and formula_type == 'shared':
                    sheet_shared_groups[shared_si].append(cell_ref)
            
            # 공유수식 그룹 메타 업데이트 (현재 시트 구간만)
            end_idx = len(self.formulas)
            if sheet_shared_groups:
                for si, cells in sheet_shared_groups.items():
                    group_size = len(cells)
                    # 그룹 bounding range 계산
                    min_row = min(self._cell_ref_to_row(c) for c in cells)
                    max_row = max(self._cell_ref_to_row(c) for c in cells)
                    min_col = min(self._cell_ref_to_col(c) for c in cells)
                    max_col = max(self._cell_ref_to_col(c) for c in cells)
                    bounding = f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{max_row}"

                    master_info = shared_masters.get(si)
                    master_cell = master_info.get("master_cell") if master_info else None
                    master_ref = master_info.get("ref") if master_info else None

                    # master_cell이 없으면(텍스트 없는 경우) 현재 시트 slice에서 formula가 있는 셀을 찾아봄
                    if not master_cell:
                        for formula_data in self.formulas[start_idx:end_idx]:
                            if (formula_data.get("formula_type") == "shared" and
                                formula_data.get("shared_si") == si and
                                formula_data.get("formula")):
                                master_cell = formula_data.get("cell_ref")
                                break

                    # 현재 시트 slice에서 메타 갱신
                    for formula_data in self.formulas[start_idx:end_idx]:
                        if (formula_data.get("formula_type") == "shared" and
                            formula_data.get("shared_si") == si):
                            formula_data["shared_master"] = master_cell
                            formula_data["shared_group_size"] = group_size
                            # ref(xml)가 있으면 보관하되, 없거나 이상하면 bounding을 사용
                            formula_data["shared_ref_range"] = master_ref or bounding

                    # 글로벌 그룹 누적(저장용)
                    self.shared_formula_groups[(sheet_name, si)].extend(cells)
        
        except Exception as e:
            logger.error(f"시트 {sheet_file} 처리 중 오류: {e}")
    
    def _check_external_ref(self, formula: str) -> bool:
        """외부 워크북 참조 확인"""
        # 패턴: [Book.xlsx]Sheet!A1 또는 '[Book.xlsx]Sheet'!A1
        pattern = r'\[.*?\]'
        return bool(re.search(pattern, formula))
    
    def _cell_ref_to_row(self, cell_ref: str) -> int:
        """셀 참조에서 행 번호 추출 (예: A1 -> 1)"""
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            return int(match.group(2))
        return 0
    
    def _cell_ref_to_col(self, cell_ref: str) -> int:
        """셀 참조에서 열 번호 추출 (예: A1 -> 1)"""
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            col_str = match.group(1)
            col = 0
            for char in col_str:
                col = col * 26 + (ord(char) - ord('A') + 1)
            return col
        return 0

    def _col_to_letters(self, col: int) -> str:
        """열 번호를 A1 표기 열 문자로 변환 (예: 1 -> A)"""
        if col <= 0:
            return "A"
        letters = ""
        while col:
            col, rem = divmod(col - 1, 26)
            letters = chr(rem + ord("A")) + letters
        return letters
    
    def save_catalog(self, output_path: str = "outputs/formula_catalog.csv"):
        """수식 카탈로그 저장"""
        df = pd.DataFrame(self.formulas)
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
        logger.info(f"수식 카탈로그 저장: {output_file} ({len(df)}개)")
        return output_file
    
    def save_groups(self, output_path: str = "outputs/formula_groups.json"):
        """공유수식 그룹 정보 저장"""
        import json
        
        groups_data = {}
        for (sheet, si), cells in self.shared_formula_groups.items():
            if len(cells) > 1:
                # range는 cells[0]:cells[-1]가 아니라 bounding box로 계산
                min_row = min(self._cell_ref_to_row(c) for c in cells)
                max_row = max(self._cell_ref_to_row(c) for c in cells)
                min_col = min(self._cell_ref_to_col(c) for c in cells)
                max_col = max(self._cell_ref_to_col(c) for c in cells)
                bounding = f"{self._col_to_letters(min_col)}{min_row}:{self._col_to_letters(max_col)}{max_row}"
                groups_data[f"{sheet}_{si}"] = {
                    "sheet": sheet,
                    "shared_si": si,
                    "cells": cells,
                    "count": len(cells),
                    "range": bounding
                }
        
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(groups_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"공유수식 그룹 저장: {output_file} ({len(groups_data)}개 그룹)")
        return output_file


def main():
    """메인 함수"""
    excel_path = r"C:\Neoprime\202511고속성장분석기(가채점)20251114 (1).xlsx"
    
    extractor = XLSXFormulaExtractor(excel_path)
    df = extractor.extract()
    
    extractor.save_catalog()
    extractor.save_groups()
    
    print(f"\n수식 추출 완료!")
    print(f"총 {len(df)}개 수식")
    if len(df) > 0:
        print(f"시트별 분포:\n{df['sheet_name'].value_counts()}")
        print(f"수식 타입별 분포:\n{df['formula_type'].value_counts()}")
    else:
        print("수식이 추출되지 않았습니다. XML 파싱 로직을 확인하세요.")


if __name__ == "__main__":
    main()
