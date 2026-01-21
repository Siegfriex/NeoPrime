"""
Excel 맥락 추출: 행/열 라벨 추출
"""

from pathlib import Path
from typing import Optional, Tuple
import logging
from openpyxl import load_workbook
from .layout_config import get_layout_config

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelContextExtractor:
    """Excel 맥락 추출기"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = None
        self._cache = {}  # (sheet, cell) -> value 캐시
    
    def open_workbook(self):
        """워크북 열기"""
        if self.wb is None:
            self.wb = load_workbook(self.excel_path, read_only=True, data_only=True)
            logger.info(f"워크북 열기: {self.excel_path}")
    
    def close_workbook(self):
        """워크북 닫기"""
        if self.wb is not None:
            self.wb.close()
            self.wb = None
    
    def get_value(self, sheet_name: str, cell_ref: str) -> Optional[str]:
        """셀 값 가져오기"""
        cache_key = (sheet_name, cell_ref)
        if cache_key in self._cache:
            return self._cache[cache_key]
        
        if self.wb is None:
            self.open_workbook()
        
        try:
            ws = self.wb[sheet_name]
            cell_value = ws[cell_ref].value
            result = str(cell_value) if cell_value is not None else None
            self._cache[cache_key] = result
            return result
        except Exception as e:
            logger.debug(f"셀 값 읽기 실패 ({sheet_name}!{cell_ref}): {e}")
            return None
    
    def parse_location(self, location: str) -> Tuple[Optional[str], Optional[str]]:
        """location 문자열(예: 'COMPUTE!G123' 또는 'RESTRICT!A10:D10')에서 시트명과 대표 셀 추출"""
        if '!' not in location:
            return None, None
        
        parts = location.split('!')
        sheet_name = parts[0]
        cell_or_range = parts[1] if len(parts) > 1 else None
        
        if not cell_or_range:
            return sheet_name, None
        
        # 범위인 경우 좌상단 셀 추출
        if ':' in cell_or_range:
            cell_ref = cell_or_range.split(':')[0]
        else:
            cell_ref = cell_or_range
        
        return sheet_name, cell_ref
    
    def extract_row_label(self, sheet_name: str, cell_ref: str) -> Optional[str]:
        """행 라벨 추출: 현재 셀에서 왼쪽으로 스캔"""
        config = get_layout_config(sheet_name)
        scan_cols = config.get('row_label_scan_cols', ['A', 'B', 'C'])
        
        # 셀 참조에서 행 번호 추출
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if not match:
            return None
        
        col_str = match.group(1)
        row_num = match.group(2)
        
        # 왼쪽으로 스캔
        col_idx = self._col_to_index(col_str)
        for col_name in scan_cols:
            scan_col_idx = self._col_to_index(col_name)
            if scan_col_idx >= col_idx:
                continue
            
            scan_cell = f"{col_name}{row_num}"
            value = self.get_value(sheet_name, scan_cell)
            if value and self._is_meaningful_label(value):
                return value.strip()
        
        return None
    
    def extract_col_header(self, sheet_name: str, cell_ref: str) -> Optional[str]:
        """열 헤더 추출: 현재 셀에서 위로 스캔"""
        config = get_layout_config(sheet_name)
        scan_rows = config.get('col_header_scan_rows', list(range(1, 6)))
        
        # 셀 참조에서 열과 행 번호 추출
        import re
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if not match:
            return None
        
        col_str = match.group(1)
        row_num = int(match.group(2))
        
        # 위로 스캔
        for scan_row in sorted(scan_rows, reverse=True):
            if scan_row >= row_num:
                continue
            
            scan_cell = f"{col_str}{scan_row}"
            value = self.get_value(sheet_name, scan_cell)
            if value and self._is_meaningful_label(value):
                return value.strip()
        
        return None
    
    def _col_to_index(self, col_str: str) -> int:
        """열 문자열(A, B, ..., Z, AA, ...)을 인덱스로 변환"""
        col = 0
        for char in col_str:
            col = col * 26 + (ord(char.upper()) - ord('A') + 1)
        return col
    
    def _is_meaningful_label(self, value: str) -> bool:
        """의미 있는 라벨인지 판단"""
        if not value or len(value.strip()) == 0:
            return False
        
        # 숫자만 있으면 제외
        if value.strip().replace('.', '').replace('-', '').isdigit():
            return False
        
        # 너무 짧거나 긴 것 제외
        if len(value.strip()) < 1 or len(value.strip()) > 100:
            return False
        
        return True
    
    def extract_context(self, location: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """location에서 시트명, 행 라벨, 열 헤더 추출"""
        sheet_name, cell_ref = self.parse_location(location)
        if not sheet_name or not cell_ref:
            return sheet_name, None, None
        
        row_label = self.extract_row_label(sheet_name, cell_ref)
        col_header = self.extract_col_header(sheet_name, cell_ref)
        
        return sheet_name, row_label, col_header


def main():
    """테스트용 메인 함수"""
    excel_path = r"C:\Neoprime\202511고속성장분석기(가채점)20251114 (1).xlsx"
    
    extractor = ExcelContextExtractor(excel_path)
    extractor.open_workbook()
    
    # 테스트
    test_locations = [
        "COMPUTE!G123",
        "RESTRICT!A10:D10",
    ]
    
    for loc in test_locations:
        sheet, row_label, col_header = extractor.extract_context(loc)
        print(f"{loc}: sheet={sheet}, row_label={row_label}, col_header={col_header}")
    
    extractor.close_workbook()


if __name__ == "__main__":
    main()
