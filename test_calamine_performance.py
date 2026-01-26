# ============================================================
# NEO GOD Ultra Calamine 엔진 성능 테스트
# test_calamine_performance.py
# Calamine 엔진 설치 및 openpyxl 대비 성능 비교
# ============================================================

import sys
import os
import unittest
import time
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime

# Windows 인코딩 설정
if sys.platform == 'win32':
    try:
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')
    except Exception:
        pass

# 엔진 가용성 확인
CALAMINE_AVAILABLE = False
OPENPYXL_AVAILABLE = False
PANDAS_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    pass

try:
    import python_calamine
    CALAMINE_AVAILABLE = True
except ImportError:
    pass

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    pass


class PerformanceTestConfig:
    """성능 테스트 설정"""

    # 테스트 파일
    SOURCE_FILE = Path(r"Y:\0126\0126\202511고속성장분석기(가채점)20251114.xlsx")

    # 테스트 반복 횟수
    NUM_ITERATIONS = 3

    # 벤치마크 시트
    BENCHMARK_SHEET = 'RAWSCORE'  # 중간 크기 시트

    # 타임아웃 (초)
    TIMEOUT = 300


class CalamineEngineTest(unittest.TestCase):
    """Calamine 엔진 테스트"""

    def test_01_calamine_installed(self):
        """Calamine 설치 확인"""
        if CALAMINE_AVAILABLE:
            import python_calamine
            print(f"\n[INFO] python-calamine 설치됨")
            print(f"  버전: {getattr(python_calamine, '__version__', 'N/A')}")
        else:
            print("\n[WARNING] python-calamine 미설치")
            print("  설치 명령: pip install python-calamine")
            self.skipTest("Calamine 미설치")

    def test_02_calamine_read_sheet_names(self):
        """Calamine으로 시트명 읽기"""
        if not CALAMINE_AVAILABLE:
            self.skipTest("Calamine 미설치")
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        from python_calamine import CalamineWorkbook

        start = time.perf_counter()
        wb = CalamineWorkbook.from_path(str(PerformanceTestConfig.SOURCE_FILE))
        sheet_names = wb.sheet_names
        elapsed = time.perf_counter() - start

        print(f"\n[INFO] Calamine 시트명 읽기: {elapsed:.3f}초")
        print(f"  시트 수: {len(sheet_names)}")

        self.assertGreater(len(sheet_names), 0)

    def test_03_calamine_read_data(self):
        """Calamine으로 데이터 읽기"""
        if not CALAMINE_AVAILABLE:
            self.skipTest("Calamine 미설치")
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        from python_calamine import CalamineWorkbook

        start = time.perf_counter()
        wb = CalamineWorkbook.from_path(str(PerformanceTestConfig.SOURCE_FILE))
        ws = wb.get_sheet_by_name(PerformanceTestConfig.BENCHMARK_SHEET)

        if ws:
            data = ws.to_python()
            elapsed = time.perf_counter() - start
            row_count = len(data) if data else 0

            print(f"\n[INFO] Calamine 데이터 읽기 ({PerformanceTestConfig.BENCHMARK_SHEET})")
            print(f"  시간: {elapsed:.3f}초")
            print(f"  행 수: {row_count:,}")

            self.assertGreater(row_count, 0)
        else:
            self.skipTest(f"시트 없음: {PerformanceTestConfig.BENCHMARK_SHEET}")


class OpenpyxlEngineTest(unittest.TestCase):
    """openpyxl 엔진 테스트"""

    def test_01_openpyxl_installed(self):
        """openpyxl 설치 확인"""
        if OPENPYXL_AVAILABLE:
            import openpyxl
            print(f"\n[INFO] openpyxl 설치됨")
            print(f"  버전: {openpyxl.__version__}")
        else:
            self.skipTest("openpyxl 미설치")

    def test_02_openpyxl_read_sheet_names(self):
        """openpyxl로 시트명 읽기"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl 미설치")
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        import openpyxl

        start = time.perf_counter()
        wb = openpyxl.load_workbook(
            str(PerformanceTestConfig.SOURCE_FILE),
            read_only=True,
            data_only=True
        )
        sheet_names = wb.sheetnames
        elapsed = time.perf_counter() - start

        print(f"\n[INFO] openpyxl 시트명 읽기: {elapsed:.3f}초")
        print(f"  시트 수: {len(sheet_names)}")

        wb.close()
        self.assertGreater(len(sheet_names), 0)

    def test_03_openpyxl_read_data(self):
        """openpyxl로 데이터 읽기"""
        if not OPENPYXL_AVAILABLE:
            self.skipTest("openpyxl 미설치")
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        import openpyxl

        start = time.perf_counter()
        wb = openpyxl.load_workbook(
            str(PerformanceTestConfig.SOURCE_FILE),
            read_only=True,
            data_only=True
        )
        ws = wb[PerformanceTestConfig.BENCHMARK_SHEET]

        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        elapsed = time.perf_counter() - start
        row_count = len(data)

        print(f"\n[INFO] openpyxl 데이터 읽기 ({PerformanceTestConfig.BENCHMARK_SHEET})")
        print(f"  시간: {elapsed:.3f}초")
        print(f"  행 수: {row_count:,}")

        wb.close()
        self.assertGreater(row_count, 0)


class PerformanceComparisonTest(unittest.TestCase):
    """Calamine vs openpyxl 성능 비교"""

    def test_01_sheet_names_comparison(self):
        """시트명 읽기 성능 비교"""
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        results = {}

        # Calamine 테스트
        if CALAMINE_AVAILABLE:
            from python_calamine import CalamineWorkbook

            times = []
            for _ in range(PerformanceTestConfig.NUM_ITERATIONS):
                start = time.perf_counter()
                wb = CalamineWorkbook.from_path(str(PerformanceTestConfig.SOURCE_FILE))
                _ = wb.sheet_names
                times.append(time.perf_counter() - start)

            results['calamine'] = sum(times) / len(times)

        # openpyxl 테스트
        if OPENPYXL_AVAILABLE:
            import openpyxl

            times = []
            for _ in range(PerformanceTestConfig.NUM_ITERATIONS):
                start = time.perf_counter()
                wb = openpyxl.load_workbook(
                    str(PerformanceTestConfig.SOURCE_FILE),
                    read_only=True,
                    data_only=True
                )
                _ = wb.sheetnames
                wb.close()
                times.append(time.perf_counter() - start)

            results['openpyxl'] = sum(times) / len(times)

        # 결과 출력
        print("\n" + "=" * 50)
        print("시트명 읽기 성능 비교")
        print("=" * 50)

        for engine, avg_time in results.items():
            print(f"  {engine}: {avg_time:.3f}초 (평균)")

        if 'calamine' in results and 'openpyxl' in results:
            speedup = results['openpyxl'] / results['calamine']
            print(f"\n  Calamine 속도 향상: {speedup:.1f}x")

            self.assertGreater(speedup, 1.0, "Calamine이 더 느림")

    def test_02_data_read_comparison(self):
        """데이터 읽기 성능 비교"""
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        results = {}
        row_counts = {}

        # Calamine 테스트
        if CALAMINE_AVAILABLE:
            from python_calamine import CalamineWorkbook

            times = []
            for _ in range(PerformanceTestConfig.NUM_ITERATIONS):
                start = time.perf_counter()
                wb = CalamineWorkbook.from_path(str(PerformanceTestConfig.SOURCE_FILE))
                ws = wb.get_sheet_by_name(PerformanceTestConfig.BENCHMARK_SHEET)
                if ws:
                    data = ws.to_python()
                    row_counts['calamine'] = len(data) if data else 0
                times.append(time.perf_counter() - start)

            results['calamine'] = sum(times) / len(times)

        # openpyxl 테스트
        if OPENPYXL_AVAILABLE:
            import openpyxl

            times = []
            for _ in range(PerformanceTestConfig.NUM_ITERATIONS):
                start = time.perf_counter()
                wb = openpyxl.load_workbook(
                    str(PerformanceTestConfig.SOURCE_FILE),
                    read_only=True,
                    data_only=True
                )
                ws = wb[PerformanceTestConfig.BENCHMARK_SHEET]
                data = list(ws.iter_rows(values_only=True))
                row_counts['openpyxl'] = len(data)
                wb.close()
                times.append(time.perf_counter() - start)

            results['openpyxl'] = sum(times) / len(times)

        # 결과 출력
        print("\n" + "=" * 50)
        print(f"데이터 읽기 성능 비교 ({PerformanceTestConfig.BENCHMARK_SHEET})")
        print("=" * 50)

        for engine, avg_time in results.items():
            rows = row_counts.get(engine, 0)
            rows_per_sec = rows / avg_time if avg_time > 0 else 0
            print(f"  {engine}: {avg_time:.3f}초 (평균), {rows:,}행, {rows_per_sec:,.0f} 행/초")

        if 'calamine' in results and 'openpyxl' in results:
            speedup = results['openpyxl'] / results['calamine']
            print(f"\n  Calamine 속도 향상: {speedup:.1f}x")

    def test_03_full_workbook_scan(self):
        """전체 워크북 스캔 성능 비교"""
        if not PerformanceTestConfig.SOURCE_FILE.exists():
            self.skipTest("소스 파일 없음")

        results = {}
        sheet_counts = {}

        # Calamine 전체 스캔
        if CALAMINE_AVAILABLE:
            from python_calamine import CalamineWorkbook

            start = time.perf_counter()
            wb = CalamineWorkbook.from_path(str(PerformanceTestConfig.SOURCE_FILE))

            total_rows = 0
            for sheet_name in wb.sheet_names:
                ws = wb.get_sheet_by_name(sheet_name)
                if ws:
                    try:
                        data = ws.to_python()
                        total_rows += len(data) if data else 0
                    except:
                        pass

            elapsed = time.perf_counter() - start
            results['calamine'] = elapsed
            sheet_counts['calamine'] = (len(wb.sheet_names), total_rows)

        # openpyxl 전체 스캔 (시간 제한 있음)
        if OPENPYXL_AVAILABLE:
            import openpyxl

            start = time.perf_counter()
            wb = openpyxl.load_workbook(
                str(PerformanceTestConfig.SOURCE_FILE),
                read_only=True,
                data_only=True
            )

            total_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    # 타임아웃 방지
                    if time.perf_counter() - start > PerformanceTestConfig.TIMEOUT:
                        break
                total_rows += row_count

                if time.perf_counter() - start > PerformanceTestConfig.TIMEOUT:
                    break

            elapsed = time.perf_counter() - start
            wb.close()
            results['openpyxl'] = elapsed
            sheet_counts['openpyxl'] = (len(wb.sheetnames) if hasattr(wb, 'sheetnames') else 0, total_rows)

        # 결과 출력
        print("\n" + "=" * 50)
        print("전체 워크북 스캔 성능 비교")
        print("=" * 50)

        for engine, elapsed in results.items():
            sheets, rows = sheet_counts.get(engine, (0, 0))
            print(f"  {engine}: {elapsed:.2f}초, {sheets}시트, {rows:,}행")

        if 'calamine' in results and 'openpyxl' in results:
            speedup = results['openpyxl'] / results['calamine']
            print(f"\n  Calamine 속도 향상: {speedup:.1f}x")


class CalamineInstallationGuide(unittest.TestCase):
    """Calamine 설치 가이드"""

    def test_installation_status(self):
        """설치 상태 및 가이드 출력"""
        print("\n" + "=" * 50)
        print("Calamine 설치 상태 및 가이드")
        print("=" * 50)

        if CALAMINE_AVAILABLE:
            print("\n[STATUS] python-calamine: 설치됨 ✅")
        else:
            print("\n[STATUS] python-calamine: 미설치 ❌")
            print("\n[설치 방법]")
            print("  pip install python-calamine")
            print("\n[기대 효과]")
            print("  - Excel 읽기 속도 10-18x 향상")
            print("  - 메모리 사용량 감소")
            print("  - 대용량 파일 처리 성능 개선")

        if OPENPYXL_AVAILABLE:
            print(f"\n[STATUS] openpyxl: 설치됨 ✅ (폴백 엔진)")
        else:
            print("\n[STATUS] openpyxl: 미설치 ❌")
            print("  pip install openpyxl")


def run_performance_tests():
    """성능 테스트 실행"""
    print("=" * 70)
    print("NEO GOD Ultra Calamine 성능 테스트")
    print("=" * 70)

    # 환경 확인
    print("\n[환경 확인]")
    print(f"  python-calamine: {'설치됨' if CALAMINE_AVAILABLE else '미설치'}")
    print(f"  openpyxl: {'설치됨' if OPENPYXL_AVAILABLE else '미설치'}")
    print(f"  pandas: {'설치됨' if PANDAS_AVAILABLE else '미설치'}")
    print(f"  소스 파일: {'존재' if PerformanceTestConfig.SOURCE_FILE.exists() else '없음'}")

    # 테스트 스위트 생성
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()

    suite.addTests(loader.loadTestsFromTestCase(CalamineEngineTest))
    suite.addTests(loader.loadTestsFromTestCase(OpenpyxlEngineTest))
    suite.addTests(loader.loadTestsFromTestCase(PerformanceComparisonTest))
    suite.addTests(loader.loadTestsFromTestCase(CalamineInstallationGuide))

    # 테스트 실행
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)

    # 결과 요약
    print("\n" + "=" * 70)
    print("성능 테스트 결과 요약")
    print("=" * 70)
    print(f"실행: {result.testsRun}")
    print(f"성공: {result.testsRun - len(result.failures) - len(result.errors) - len(result.skipped)}")
    print(f"실패: {len(result.failures)}")
    print(f"에러: {len(result.errors)}")
    print(f"스킵: {len(result.skipped)}")

    return result.wasSuccessful()


if __name__ == '__main__':
    success = run_performance_tests()
    sys.exit(0 if success else 1)
