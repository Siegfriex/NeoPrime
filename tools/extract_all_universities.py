"""
전체 550개 대학 환산점수 추출 스크립트

Phase 2: Phase 1의 18개 PoC → 550개 전체 대학으로 확장

사용법:
    python tools/extract_all_universities.py

출력:
    theory_engine/weights/subject3_conversions_full.json

주의:
    - Windows + Excel 설치 환경에서만 동작
    - xlwings/openpyxl 필요
    - 추출 시에만 실행 (런타임에는 JSON만 사용)
"""

import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# 프로젝트 루트 추가
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


class UniversityExtractor:
    """전체 대학 환산점수 추출기

    SUBJECT3 시트에서 모든 대학의 환산점수 테이블을 추출합니다.
    """

    # 과목 행 매핑 (SUBJECT3 시트 기준)
    SUBJECT_ROWS = {
        "국어": list(range(5, 204)),      # 표준점수 200~5 (실제 행 번호)
        "수학": list(range(207, 356)),    # 표준점수 200~51 (추정)
        "영어": list(range(359, 368)),    # 등급 1~9
        "탐구": list(range(371, 471)),    # 탐구 환산점수
        "한국사": list(range(474, 483)),  # 등급 1~9
    }

    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.ws_subject3 = None

    def extract_all(self, use_xlwings: bool = True) -> Dict:
        """전체 대학 환산점수 추출

        Args:
            use_xlwings: True면 xlwings (COM), False면 openpyxl

        Returns:
            완전한 환산점수 데이터 딕셔너리
        """
        if use_xlwings:
            return self._extract_with_xlwings()
        else:
            return self._extract_with_openpyxl()

    def _extract_with_xlwings(self) -> Dict:
        """xlwings로 추출 (COM 기반 - 수식 계산 결과 획득)"""
        try:
            import xlwings as xw
        except ImportError:
            logger.error("xlwings가 설치되지 않았습니다: pip install xlwings")
            raise

        logger.info(f"xlwings로 추출 시작: {self.excel_path}")

        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False

        result = {
            "metadata": {
                "source_excel": str(self.excel_path.name),
                "extraction_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "extraction_method": "xlwings_com",
                "total_rows": 0,
                "total_cols": 0,
                "total_universities": 0,
            },
            "university_mapping": {},
            "conversion_table": {},
        }

        try:
            wb = app.books.open(str(self.excel_path))

            # SUBJECT3 시트 찾기
            ws_subject3 = None
            for sheet in wb.sheets:
                if 'SUBJECT3' in sheet.name.upper() or sheet.name == 'SUBJECT3':
                    ws_subject3 = sheet
                    break

            if ws_subject3 is None:
                # 시트 이름으로 찾기 실패시 인덱스로 시도
                try:
                    ws_subject3 = wb.sheets['SUBJECT3']
                except:
                    logger.error("SUBJECT3 시트를 찾을 수 없습니다.")
                    raise ValueError("SUBJECT3 시트 없음")

            logger.info(f"SUBJECT3 시트 발견: {ws_subject3.name}")

            # 데이터 범위 확인
            used_range = ws_subject3.used_range
            max_row = used_range.last_cell.row
            max_col = used_range.last_cell.column

            logger.info(f"데이터 범위: {max_row}행 x {max_col}열")
            result["metadata"]["total_rows"] = max_row
            result["metadata"]["total_cols"] = max_col

            # 1. 대학 매핑 추출 (1~4행: 대학명, 학과명, 계열, 필수과목)
            logger.info("대학 매핑 추출 중...")
            university_mapping = self._extract_university_mapping(ws_subject3, max_col)
            result["university_mapping"] = university_mapping
            result["metadata"]["total_universities"] = len(university_mapping)

            # 2. 환산점수 테이블 추출 (모든 대학에 대해)
            logger.info(f"환산점수 테이블 추출 중... ({len(university_mapping)}개 대학)")
            conversion_table = self._extract_conversion_table(
                ws_subject3, university_mapping, max_row
            )
            result["conversion_table"] = conversion_table

            logger.info(f"추출 완료: {len(conversion_table)}개 대학 환산점수")

        except Exception as e:
            logger.error(f"추출 중 오류: {e}")
            import traceback
            traceback.print_exc()
            raise

        finally:
            try:
                wb.close()
            except:
                pass
            app.quit()

        return result

    def _extract_with_openpyxl(self) -> Dict:
        """openpyxl로 추출 (수식은 data_only=True로 값만)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl이 설치되지 않았습니다: pip install openpyxl")
            raise

        logger.info(f"openpyxl로 추출 시작: {self.excel_path}")

        # data_only=True로 수식 대신 계산된 값 로드
        wb = load_workbook(self.excel_path, data_only=True)

        result = {
            "metadata": {
                "source_excel": str(self.excel_path.name),
                "extraction_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "extraction_method": "openpyxl_data_only",
                "total_rows": 0,
                "total_cols": 0,
                "total_universities": 0,
            },
            "university_mapping": {},
            "conversion_table": {},
        }

        try:
            ws_subject3 = wb['SUBJECT3']

            max_row = ws_subject3.max_row
            max_col = ws_subject3.max_column

            logger.info(f"데이터 범위: {max_row}행 x {max_col}열")
            result["metadata"]["total_rows"] = max_row
            result["metadata"]["total_cols"] = max_col

            # openpyxl용 추출 로직
            # (xlwings와 유사하지만 셀 접근 방식이 다름)
            # TODO: 필요시 구현

        finally:
            wb.close()

        return result

    def _extract_university_mapping(
        self,
        ws,
        max_col: int
    ) -> Dict[str, Dict]:
        """대학 매핑 추출

        행 1: 대학명
        행 2: 학과명
        행 3: 계열/method
        행 4: 필수과목 문자열
        """
        mapping = {}

        # K열(11)부터 시작 (A~J는 점수 입력 영역)
        start_col = 11  # K

        for col_idx in range(start_col, max_col + 1):
            col_letter = self._col_idx_to_letter(col_idx)

            try:
                university = ws.range(f'{col_letter}1').value
                department = ws.range(f'{col_letter}2').value
                method = ws.range(f'{col_letter}3').value
                required = ws.range(f'{col_letter}4').value

                # 유효한 대학만 추가
                if university and department:
                    mapping[col_letter] = {
                        "index": col_idx,
                        "university": str(university).strip(),
                        "department": str(department).strip(),
                        "method": str(method).strip() if method else "",
                        "required_subjects": str(required).strip() if required else "",
                    }

            except Exception as e:
                logger.debug(f"컬럼 {col_letter} 처리 중 오류: {e}")
                continue

        logger.info(f"대학 매핑 추출: {len(mapping)}개")
        return mapping

    def _extract_conversion_table(
        self,
        ws,
        university_mapping: Dict[str, Dict],
        max_row: int
    ) -> Dict[str, Dict]:
        """환산점수 테이블 추출

        각 대학별로 과목-점수 → 환산점수 매핑 추출
        """
        conversion_table = {}

        # 점수 범위 정의 (SUBJECT3 시트 구조 기준)
        SCORE_RANGES = {
            "국어": {
                "start_row": 5,
                "end_row": 204,  # 200행
                "score_col": "A",  # 점수가 있는 컬럼
                "score_range": (5, 200),  # 표준점수 범위
            },
            "수학": {
                "start_row": 207,
                "end_row": 356,
                "score_col": "A",
                "score_range": (51, 200),
            },
            "영어": {
                "start_row": 359,
                "end_row": 367,
                "score_col": "A",
                "score_range": (1, 9),  # 등급
            },
            # 탐구는 과목명별로 별도 처리 필요
            "탐구1": {
                "start_row": 371,
                "end_row": 470,
                "score_col": "A",
                "score_range": (1, 100),
            },
            "탐구2": {
                "start_row": 371,
                "end_row": 470,
                "score_col": "A",
                "score_range": (1, 100),
            },
            "한국사": {
                "start_row": 474,
                "end_row": 482,
                "score_col": "A",
                "score_range": (1, 9),  # 등급
            },
        }

        total = len(university_mapping)
        processed = 0

        for col_letter, univ_info in university_mapping.items():
            univ_name = univ_info["university"]
            dept_name = univ_info["department"]
            key = f"{univ_name}_{dept_name}"

            conversions = {}

            # 각 과목별 환산점수 추출
            for subject, range_info in SCORE_RANGES.items():
                start_row = range_info["start_row"]
                end_row = range_info["end_row"]

                for row in range(start_row, min(end_row + 1, max_row + 1)):
                    try:
                        # A열에서 원점수/등급 읽기
                        score = ws.range(f'A{row}').value
                        if score is None:
                            continue

                        # 해당 대학 컬럼에서 환산점수 읽기
                        converted = ws.range(f'{col_letter}{row}').value
                        if converted is not None:
                            try:
                                converted_float = float(converted)
                                # 유효한 값만 저장
                                if converted_float >= 0:
                                    score_key = f"{subject}-{int(score)}"
                                    conversions[score_key] = converted_float
                            except (ValueError, TypeError):
                                pass

                    except Exception as e:
                        logger.debug(f"셀 {col_letter}{row} 처리 중 오류: {e}")
                        continue

            if conversions:
                conversion_table[key] = {
                    "university": univ_name,
                    "department": dept_name,
                    "method": univ_info.get("method", ""),
                    "required_subjects": univ_info.get("required_subjects", ""),
                    "conversions": conversions,
                    "conversion_count": len(conversions),
                }

            processed += 1
            if processed % 50 == 0:
                logger.info(f"진행률: {processed}/{total} ({processed/total*100:.1f}%)")

        return conversion_table

    def _col_idx_to_letter(self, col_idx: int) -> str:
        """열 인덱스 → 열 문자 변환 (1=A, 27=AA, ...)"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def save_result(self, result: Dict, output_path: str = None):
        """결과 저장"""
        if output_path is None:
            output_path = PROJECT_ROOT / "theory_engine" / "weights" / "subject3_conversions_full.json"
        else:
            output_path = Path(output_path)

        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        logger.info(f"저장 완료: {output_path}")
        logger.info(f"  - 총 대학/학과: {result['metadata']['total_universities']}개")
        logger.info(f"  - 환산점수 테이블: {len(result['conversion_table'])}개")

        return output_path


def main():
    """메인 함수"""
    excel_path = PROJECT_ROOT / "202511고속성장분석기(가채점)20251114 (1).xlsx"

    if not excel_path.exists():
        logger.error(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")
        return

    extractor = UniversityExtractor(str(excel_path))

    try:
        # xlwings로 추출 (COM 기반)
        result = extractor.extract_all(use_xlwings=True)

        # 결과 저장
        output_path = extractor.save_result(result)

        # 검증
        print("\n" + "=" * 60)
        print("추출 완료 요약")
        print("=" * 60)
        print(f"출력 파일: {output_path}")
        print(f"총 대학/학과: {result['metadata']['total_universities']}개")
        print(f"환산점수 테이블: {len(result['conversion_table'])}개")

        if len(result['conversion_table']) >= 500:
            print("\n SUCCESS: 500개 이상 대학 추출 완료!")
        else:
            print(f"\n WARNING: {len(result['conversion_table'])}개만 추출됨 (목표: 500+)")

    except Exception as e:
        logger.error(f"추출 실패: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
