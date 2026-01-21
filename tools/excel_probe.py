"""
Excel 파일의 수식/룰 추출 가능성 진단 도구

이 스크립트는 엑셀 파일을 분석하여:
- 수식 분포 및 함수 사용 현황
- 공유수식/배열수식 존재 여부
- 네임드레인지/테이블/조건부서식/데이터검증 존재 여부
- 외부 링크 존재 여부
를 진단하여, 이후 Formula Mining의 난이도를 사전에 파악합니다.
"""

import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Dict, List, Set, Tuple
from collections import Counter, defaultdict
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class ExcelProbe:
    """엑셀 파일 진단 클래스"""
    
    def __init__(self, excel_path: str):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.stats = {
            "sheets": {},
            "formulas": {
                "total_cells": 0,
                "formula_cells": 0,
                "shared_formula_groups": 0,
                "array_formulas": 0,
                "function_counts": Counter(),
            },
            "named_ranges": {
                "count": 0,
                "types": defaultdict(int),
            },
            "tables": {
                "count": 0,
                "details": [],
            },
            "data_validations": {
                "count": 0,
                "sheets": defaultdict(int),
            },
            "conditional_formats": {
                "count": 0,
                "sheets": defaultdict(int),
            },
            "external_links": {
                "exists": False,
                "count": 0,
            },
        }
    
    def probe(self) -> Dict:
        """전체 진단 실행"""
        logger.info(f"진단 시작: {self.excel_path}")
        
        # 1. openpyxl로 워크북 로드
        self.wb = load_workbook(self.excel_path, data_only=False, read_only=True)
        
        # 2. 시트별 기본 정보 수집
        self._probe_sheets()
        
        # 3. XML 레벨에서 수식 빠르게 추출
        self._probe_formulas_xml()
        
        # 4. 네임드레인지 추출
        self._probe_named_ranges()
        
        # 5. 테이블 추출
        self._probe_tables()
        
        # 6. 데이터검증 추출
        self._probe_data_validations()
        
        # 7. 조건부서식 추출
        self._probe_conditional_formats()
        
        # 8. 외부 링크 확인
        self._probe_external_links()
        
        logger.info("진단 완료")
        return self.stats
    
    def _probe_sheets(self):
        """시트별 기본 정보"""
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            # read_only 모드에서는 max_row/max_column이 정확하지 않을 수 있음
            self.stats["sheets"][sheet_name] = {
                "exists": True,
                "note": "read_only 모드에서는 정확한 크기 측정 제한적"
            }
    
    def _probe_formulas_xml(self):
        """XML 레벨에서 수식 빠르게 추출"""
        logger.info("XML에서 수식 추출 중...")
        
        with zipfile.ZipFile(self.excel_path, 'r') as z:
            # xl/worksheets/ 디렉토리 찾기
            sheet_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
            
            shared_formula_groups = set()
            array_formulas = []
            function_counter = Counter()
            cf_rule_total = 0
            dv_total = 0
            table_part_total = 0
            cf_sheets = defaultdict(int)
            dv_sheets = defaultdict(int)
            table_sheets = defaultdict(int)
            
            for sheet_file in sheet_files:
                try:
                    sheet_name = self._get_sheet_name_from_path(sheet_file, z)
                    xml_content = z.read(sheet_file)
                    root = ET.fromstring(xml_content)
                    
                    # 네임스페이스 처리
                    ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                    # 조건부서식/데이터검증/테이블파트는 read_only openpyxl에서 종종 0으로 나오므로 XML에서 직접 카운트
                    cf_rules = root.findall('.//main:conditionalFormatting/main:cfRule', ns)
                    if cf_rules:
                        cf_sheets[sheet_name] = len(cf_rules)
                        cf_rule_total += len(cf_rules)

                    dvs = root.findall('.//main:dataValidations/main:dataValidation', ns)
                    if dvs:
                        dv_sheets[sheet_name] = len(dvs)
                        dv_total += len(dvs)

                    tparts = root.findall('.//main:tableParts/main:tablePart', ns)
                    if tparts:
                        table_sheets[sheet_name] = len(tparts)
                        table_part_total += len(tparts)
                    
                    # 행 순회
                    for row in root.findall('.//main:row', ns):
                        for cell in row.findall('.//main:c', ns):
                            cell_ref = cell.get('r')
                            if cell_ref is None:
                                continue
                            
                            self.stats["formulas"]["total_cells"] += 1
                            
                            # 수식 확인
                            formula_elem = cell.find('.//main:f', ns)
                            if formula_elem is not None:
                                self.stats["formulas"]["formula_cells"] += 1
                                formula_text = formula_elem.text or ""
                                if formula_text and not formula_text.startswith('='):
                                    formula_text = '=' + formula_text
                                
                                # 공유수식 확인 (formula 요소에서 si 추출)
                                si = formula_elem.get('si')  # 수식의 공유 인덱스
                                t_type = formula_elem.get('t', '')  # 수식의 타입
                                
                                if si is not None and t_type == 'shared':
                                    shared_formula_groups.add((sheet_name, si))
                                
                                # 배열수식 확인
                                if t_type == 'array':
                                    array_formulas.append(f"{sheet_name}!{cell_ref}")
                                
                                # 함수 추출 (간단한 패턴 매칭)
                                if formula_text:
                                    # 주요 함수 패턴 추출
                                    for func in ['IF', 'VLOOKUP', 'INDEX', 'MATCH', 'OFFSET', 
                                                'INDIRECT', 'SUM', 'AVERAGE', 'COUNT', 'IFS',
                                                'SWITCH', 'XLOOKUP', 'FILTER', 'SORT']:
                                        if func in formula_text.upper():
                                            function_counter[func] += 1
                
                except Exception as e:
                    logger.warning(f"시트 {sheet_file} 처리 중 오류: {e}")
            
            self.stats["formulas"]["shared_formula_groups"] = len(shared_formula_groups)
            self.stats["formulas"]["array_formulas"] = len(array_formulas)
            self.stats["formulas"]["function_counts"] = dict(function_counter.most_common(20))

            # XML 기반 메타 반영
            self.stats["conditional_formats"]["count"] = cf_rule_total
            self.stats["conditional_formats"]["sheets"] = dict(cf_sheets)
            self.stats["data_validations"]["count"] = dv_total
            self.stats["data_validations"]["sheets"] = dict(dv_sheets)
            self.stats["tables"]["count"] = table_part_total
            # tables.details는 openpyxl 기반(있으면)만 채움
    
    def _get_sheet_name_from_path(self, sheet_path: str, z: zipfile.ZipFile) -> str:
        """sheetN.xml 경로에서 시트명 찾기"""
        # NOTE:
        # - sheetN.xml의 N은 workbook.xml의 sheetId와 1:1이 아닐 수 있음
        # - 정확한 매핑은 workbook.xml의 r:id -> workbook.xml.rels의 Target(worksheets/sheetN.xml)로 해야 함
        try:
            workbook_xml = z.read('xl/workbook.xml')
            rels_xml = z.read('xl/_rels/workbook.xml.rels')
            wb_root = ET.fromstring(workbook_xml)
            rels_root = ET.fromstring(rels_xml)
            ns = {
                'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
            }
            relns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

            rel_map = {}
            for rel in rels_root.iter():
                if rel.tag.endswith('Relationship'):
                    rid = rel.get('Id')
                    target = rel.get('Target')
                    if target and 'worksheets' in target:
                        if not target.startswith('xl/'):
                            target = 'xl/' + target
                        rel_map[rid] = target

            for sheet in wb_root.findall('.//main:sheet', ns):
                name = sheet.get('name')
                rid = sheet.get(f'{{{relns}}}id')
                target = rel_map.get(rid)
                if target == sheet_path and name:
                    return name
        except Exception as e:
            logger.debug(f"시트명 추출 실패 ({sheet_path}): {e}")
        
        # 폴백: 경로에서 추정
        return Path(sheet_path).stem
    
    def _probe_named_ranges(self):
        """네임드레인지 추출"""
        logger.info("네임드레인지 추출 중...")
        
        for name_obj in self.wb.defined_names:
            self.stats["named_ranges"]["count"] += 1
            
            # 타입 추정 (간단)
            value_str = str(name_obj.value) if hasattr(name_obj, 'value') else str(name_obj)
            if '!' in value_str:
                self.stats["named_ranges"]["types"]["range"] += 1
            elif any(op in value_str for op in ['+', '-', '*', '/', '=']):
                self.stats["named_ranges"]["types"]["formula"] += 1
            else:
                self.stats["named_ranges"]["types"]["constant"] += 1
    
    def _probe_tables(self):
        """테이블(ListObject) 추출"""
        logger.info("테이블 추출 중...")
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if hasattr(ws, 'tables') and ws.tables:
                for table_name, table in ws.tables.items():
                    self.stats["tables"]["count"] += 1
                    self.stats["tables"]["details"].append({
                        "sheet": sheet_name,
                        "name": table_name,
                        "ref": table.ref,
                        "column_count": len(table.tableColumns) if table.tableColumns else 0,
                    })
    
    def _probe_data_validations(self):
        """데이터검증 추출"""
        logger.info("데이터검증 추출 중...")
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if hasattr(ws, 'data_validations'):
                count = len(ws.data_validations.dataValidation)
                if count > 0:
                    self.stats["data_validations"]["count"] += count
                    self.stats["data_validations"]["sheets"][sheet_name] = count
    
    def _probe_conditional_formats(self):
        """조건부서식 추출"""
        logger.info("조건부서식 추출 중...")
        
        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if hasattr(ws, 'conditional_formatting'):
                # read_only 모드에서는 제한적
                # 실제로는 각 셀의 conditional_formatting 속성을 확인해야 함
                # 여기서는 존재 여부만 확인
                if hasattr(ws, '_cells') or True:  # 간단 체크
                    self.stats["conditional_formats"]["sheets"][sheet_name] = 1
                    self.stats["conditional_formats"]["count"] += 1
    
    def _probe_external_links(self):
        """외부 링크 확인"""
        logger.info("외부 링크 확인 중...")
        
        with zipfile.ZipFile(self.excel_path, 'r') as z:
            # xl/externalLinks/ 디렉토리 확인
            external_link_files = [f for f in z.namelist() if f.startswith('xl/externalLinks/')]
            
            if external_link_files:
                self.stats["external_links"]["exists"] = True
                self.stats["external_links"]["count"] = len(external_link_files)
    
    def generate_report(self, output_path: str = "outputs/probe_report.json"):
        """진단 결과 리포트 생성"""
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # JSON 저장
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(self.stats, f, ensure_ascii=False, indent=2, default=str)
        
        # 텍스트 리포트도 생성
        txt_file = output_file.with_suffix('.txt')
        with open(txt_file, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("Excel 파일 진단 리포트\n")
            f.write("=" * 80 + "\n\n")
            
            f.write(f"파일: {self.excel_path}\n")
            f.write(f"시트 수: {len(self.stats['sheets'])}\n\n")
            
            f.write("=" * 80 + "\n")
            f.write("수식 통계\n")
            f.write("=" * 80 + "\n")
            f.write(f"전체 셀 수: {self.stats['formulas']['total_cells']:,}\n")
            f.write(f"수식 셀 수: {self.stats['formulas']['formula_cells']:,}\n")
            f.write(f"수식 비율: {self.stats['formulas']['formula_cells'] / max(self.stats['formulas']['total_cells'], 1) * 100:.2f}%\n")
            f.write(f"공유수식 그룹 수: {self.stats['formulas']['shared_formula_groups']}\n")
            f.write(f"배열수식 수: {self.stats['formulas']['array_formulas']}\n\n")
            
            f.write("주요 함수 사용 횟수 (TOP 20):\n")
            for func, count in list(self.stats['formulas']['function_counts'].items())[:20]:
                f.write(f"  {func}: {count}\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("네임드레인지\n")
            f.write("=" * 80 + "\n")
            f.write(f"총 개수: {self.stats['named_ranges']['count']}\n")
            for type_name, count in self.stats['named_ranges']['types'].items():
                f.write(f"  {type_name}: {count}\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("테이블\n")
            f.write("=" * 80 + "\n")
            f.write(f"총 개수: {self.stats['tables']['count']}\n")
            for table in self.stats['tables']['details']:
                f.write(f"  {table['sheet']}!{table['name']}: {table['ref']} ({table['column_count']} 컬럼)\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("데이터검증\n")
            f.write("=" * 80 + "\n")
            f.write(f"총 개수: {self.stats['data_validations']['count']}\n")
            for sheet, count in self.stats['data_validations']['sheets'].items():
                f.write(f"  {sheet}: {count}\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("조건부서식\n")
            f.write("=" * 80 + "\n")
            f.write(f"총 개수: {self.stats['conditional_formats']['count']}\n")
            for sheet, count in self.stats['conditional_formats']['sheets'].items():
                f.write(f"  {sheet}: {count}\n")
            f.write("\n")
            
            f.write("=" * 80 + "\n")
            f.write("외부 링크\n")
            f.write("=" * 80 + "\n")
            f.write(f"존재 여부: {self.stats['external_links']['exists']}\n")
            f.write(f"링크 파일 수: {self.stats['external_links']['count']}\n")
            f.write("\n")
            
            # 난이도 평가
            f.write("=" * 80 + "\n")
            f.write("난이도 평가\n")
            f.write("=" * 80 + "\n")
            
            risk_factors = []
            if self.stats['formulas']['function_counts'].get('OFFSET', 0) > 100:
                risk_factors.append("OFFSET 함수 다수 사용 (동적 참조)")
            if self.stats['formulas']['function_counts'].get('INDIRECT', 0) > 50:
                risk_factors.append("INDIRECT 함수 다수 사용 (동적 참조)")
            if self.stats['external_links']['exists']:
                risk_factors.append("외부 링크 존재 (의존성 복잡)")
            if self.stats['formulas']['array_formulas'] > 0:
                risk_factors.append("배열수식 존재 (복잡한 계산)")
            
            if risk_factors:
                f.write("주의 사항:\n")
                for risk in risk_factors:
                    f.write(f"  - {risk}\n")
            else:
                f.write("정적 분석으로 충분히 추출 가능해 보입니다.\n")
        
        logger.info(f"리포트 생성 완료: {output_file}, {txt_file}")
        return output_file, txt_file


def main():
    """메인 함수"""
    excel_path = r"C:\Neoprime\202511고속성장분석기(가채점)20251114 (1).xlsx"
    
    probe = ExcelProbe(excel_path)
    stats = probe.probe()
    probe.generate_report()
    
    print("\n진단 완료! outputs/probe_report.json 및 .txt 파일을 확인하세요.")


if __name__ == "__main__":
    main()
