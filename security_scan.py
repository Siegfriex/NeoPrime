#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 파일 보안 진단 스크립트
- 파일 암호 확인 (msoffcrypto-tool)
- 시트 보호 확인 (openpyxl)
- 숨김 시트 확인 (openpyxl)
"""

import sys
import os
from pathlib import Path
from datetime import datetime

# 출력 인코딩 설정 (Windows 한글 출력을 위해)
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

# 타겟 파일 경로 하드코딩
TARGET_FILE = r"Y:\0126\0126\202511고속성장분석기(가채점)20251114.xlsx"

def check_file_encryption(file_path):
    """
    Level 1: 파일 암호 확인
    msoffcrypto-tool을 사용하여 파일 암호화 여부 확인
    """
    print("\n[Level 1] 파일 암호 확인 중...")
    print("-" * 60)
    
    try:
        import msoffcrypto
        
        with open(file_path, "rb") as f:
            try:
                office_file = msoffcrypto.OfficeFile(f)
                # 비밀번호 없이 키 로드 시도
                office_file.load_key()
                print("  [결과] 파일 암호화되지 않음 - 비밀번호 없이 접근 가능")
                return {"encrypted": False, "password_required": False, "status": "SUCCESS"}
            except Exception as e:
                error_msg = str(e).lower()
                if "password" in error_msg or "key" in error_msg or "no key" in error_msg:
                    print("  [결과] 파일이 암호화되어 있음 - 비밀번호 필요")
                    print(f"  [상세] {str(e)}")
                    # openpyxl로 열 수 있는지 확인 (일부 암호화는 openpyxl이 우회 가능)
                    try:
                        from openpyxl import load_workbook
                        test_wb = load_workbook(file_path, read_only=True)
                        test_wb.close()
                        print("  [참고] openpyxl로는 파일을 열 수 있음 (암호화가 완전하지 않거나 특정 방식)")
                        return {"encrypted": True, "password_required": True, "status": "WARNING", "error": str(e), "openpyxl_readable": True}
                    except:
                        return {"encrypted": True, "password_required": True, "status": "CRITICAL", "error": str(e), "openpyxl_readable": False}
                else:
                    print(f"  [결과] 파일 확인 중 오류 발생: {str(e)}")
                    return {"encrypted": None, "password_required": None, "status": "ERROR", "error": str(e)}
                    
    except ImportError:
        print("  [경고] msoffcrypto-tool이 설치되지 않음")
        print("  [조치] pip install msoffcrypto-tool 실행 필요")
        print("  [임시] 파일 암호 확인을 건너뜀 (openpyxl로 시트 보호만 확인)")
        return {"encrypted": None, "password_required": None, "status": "SKIP", "reason": "msoffcrypto-tool not installed"}
    except Exception as e:
        print(f"  [오류] 파일 암호 확인 중 예외 발생: {str(e)}")
        return {"encrypted": None, "password_required": None, "status": "ERROR", "error": str(e)}


def check_sheet_protection(file_path):
    """
    Level 2: 시트 보호 확인
    openpyxl을 사용하여 시트 보호 상태 확인
    """
    print("\n[Level 2] 시트 보호 확인 중...")
    print("-" * 60)
    
    protected_sheets = []
    unprotected_sheets = []
    
    try:
        from openpyxl import load_workbook
        
        print("  [진행] 워크북 열기 중...")
        wb = load_workbook(file_path, read_only=False, data_only=False)
        
        print(f"  [정보] 총 시트 수: {len(wb.sheetnames)}")
        
        for sheet_name in wb.sheetnames:
            try:
                sheet = wb[sheet_name]
                is_protected = sheet.protection.sheet if hasattr(sheet.protection, 'sheet') else False
                
                if is_protected:
                    protected_sheets.append(sheet_name)
                    print(f"    [보호됨] '{sheet_name}'")
                else:
                    unprotected_sheets.append(sheet_name)
                    
            except Exception as e:
                print(f"    [오류] 시트 '{sheet_name}' 확인 중 오류: {str(e)}")
        
        wb.close()
        
        print(f"\n  [요약] 보호된 시트: {len(protected_sheets)}개")
        print(f"  [요약] 보호되지 않은 시트: {len(unprotected_sheets)}개")
        
        return {
            "protected_sheets": protected_sheets,
            "unprotected_sheets": unprotected_sheets,
            "total_sheets": len(wb.sheetnames),
            "status": "SUCCESS"
        }
        
    except ImportError:
        print("  [오류] openpyxl이 설치되지 않음")
        return {"status": "ERROR", "error": "openpyxl not installed"}
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            print(f"  [결과] 파일이 암호화되어 있어 열 수 없음: {str(e)}")
            return {"status": "CRITICAL", "error": str(e), "reason": "File is encrypted"}
        else:
            print(f"  [오류] 워크북 열기 실패: {str(e)}")
            return {"status": "ERROR", "error": str(e)}


def check_hidden_sheets(file_path):
    """
    Level 3: 숨김 시트 확인
    openpyxl을 사용하여 숨겨진 시트 확인
    """
    print("\n[Level 3] 숨김 시트 확인 중...")
    print("-" * 60)
    
    visible_sheets = []
    hidden_sheets = []
    very_hidden_sheets = []
    
    try:
        from openpyxl import load_workbook
        
        print("  [진행] 워크북 열기 중...")
        wb = load_workbook(file_path, read_only=False, data_only=False)
        
        for sheet_name in wb.sheetnames:
            try:
                sheet = wb[sheet_name]
                sheet_state = sheet.sheet_state if hasattr(sheet, 'sheet_state') else 'visible'
                
                if sheet_state == 'visible':
                    visible_sheets.append(sheet_name)
                elif sheet_state == 'hidden':
                    hidden_sheets.append(sheet_name)
                    print(f"    [숨김] '{sheet_name}' (hidden)")
                elif sheet_state == 'veryHidden':
                    very_hidden_sheets.append(sheet_name)
                    print(f"    [완전숨김] '{sheet_name}' (veryHidden)")
                else:
                    print(f"    [알 수 없음] '{sheet_name}' (state: {sheet_state})")
                    
            except Exception as e:
                print(f"    [오류] 시트 '{sheet_name}' 상태 확인 중 오류: {str(e)}")
        
        wb.close()
        
        print(f"\n  [요약] 보이는 시트: {len(visible_sheets)}개")
        print(f"  [요약] 숨김 시트: {len(hidden_sheets)}개")
        print(f"  [요약] 완전숨김 시트: {len(very_hidden_sheets)}개")
        
        return {
            "visible_sheets": visible_sheets,
            "hidden_sheets": hidden_sheets,
            "very_hidden_sheets": very_hidden_sheets,
            "status": "SUCCESS"
        }
        
    except ImportError:
        print("  [오류] openpyxl이 설치되지 않음")
        return {"status": "ERROR", "error": "openpyxl not installed"}
    except Exception as e:
        error_msg = str(e).lower()
        if "password" in error_msg or "encrypted" in error_msg:
            print(f"  [결과] 파일이 암호화되어 있어 열 수 없음: {str(e)}")
            return {"status": "CRITICAL", "error": str(e), "reason": "File is encrypted"}
        else:
            print(f"  [오류] 워크북 열기 실패: {str(e)}")
            return {"status": "ERROR", "error": str(e)}


def get_file_metadata(file_path):
    """파일 메타데이터 수집"""
    try:
        path_obj = Path(file_path)
        if path_obj.exists():
            stat = path_obj.stat()
            return {
                "exists": True,
                "size_bytes": stat.st_size,
                "size_mb": round(stat.st_size / (1024 * 1024), 2),
                "modified_time": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S")
            }
        else:
            return {"exists": False}
    except Exception as e:
        return {"exists": False, "error": str(e)}


def generate_report(encryption_result, protection_result, hidden_result, metadata):
    """
    보안 진단 보고서 생성
    """
    print("\n" + "=" * 60)
    print("## [보안 진단 보고서]")
    print("=" * 60)
    
    # 파일 정보
    print(f"\n[파일 정보]")
    print(f"  경로: {TARGET_FILE}")
    if metadata.get("exists"):
        print(f"  크기: {metadata.get('size_mb', 0)} MB ({metadata.get('size_bytes', 0):,} bytes)")
        print(f"  수정일시: {metadata.get('modified_time', 'N/A')}")
    else:
        print(f"  [오류] 파일을 찾을 수 없습니다.")
        return
    
    # 1. 파일 접근성
    print(f"\n1. **파일 접근성**: ", end="")
    if encryption_result.get("status") == "CRITICAL":
        print("[실패] (비밀번호 필요)")
        print(f"   - 파일이 암호화되어 있어 비밀번호 없이는 접근 불가능합니다.")
        print(f"   - 오류 메시지: {encryption_result.get('error', 'N/A')}")
    elif encryption_result.get("status") == "WARNING":
        print("[부분 제한] (암호화 감지됨, 하지만 openpyxl로 읽기 가능)")
        print(f"   - 파일이 암호화되어 있지만, openpyxl/Pandas로는 읽을 수 있습니다.")
        print(f"   - msoffcrypto-tool 오류: {encryption_result.get('error', 'N/A')}")
        print(f"   - 권장: Pandas로 직접 읽기 시도 가능")
    elif encryption_result.get("status") == "SUCCESS":
        print("[성공] (비밀번호 없음)")
        print(f"   - 파일이 암호화되지 않아 비밀번호 없이 접근 가능합니다.")
    elif encryption_result.get("status") == "SKIP":
        print("[확인 불가] (msoffcrypto-tool 미설치)")
        print(f"   - 파일 암호 확인을 건너뜀: {encryption_result.get('reason', 'N/A')}")
    else:
        print("[확인 불가] (오류 발생)")
        if encryption_result.get("error"):
            print(f"   - 오류: {encryption_result.get('error')}")
    
    # 2. 시트 보호 현황
    print(f"\n2. **시트 보호 현황**:")
    if protection_result.get("status") == "SUCCESS":
        protected_count = len(protection_result.get("protected_sheets", []))
        unprotected_count = len(protection_result.get("unprotected_sheets", []))
        total_count = protection_result.get("total_sheets", 0)
        
        print(f"   - 보호된 시트: {protected_count}개", end="")
        if protected_count > 0:
            print(f" ({', '.join(protection_result.get('protected_sheets', [])[:5])}", end="")
            if protected_count > 5:
                print(f" 외 {protected_count - 5}개...", end="")
            print(")")
        else:
            print()
        
        print(f"   - 보호되지 않은 시트: {unprotected_count}개")
        print(f"   - 데이터 추출 가능 여부: [가능]")
        print(f"     (Pandas는 시트 보호를 무시하고 읽을 수 있습니다)")
    elif protection_result.get("status") == "CRITICAL":
        print(f"   - [확인 불가] 파일이 암호화되어 있어 시트 보호 상태를 확인할 수 없습니다.")
        print(f"   - 오류: {protection_result.get('error', 'N/A')}")
    else:
        print(f"   - [확인 불가] 오류 발생: {protection_result.get('error', 'N/A')}")
    
    # 3. 숨겨진 시트
    print(f"\n3. **숨겨진 시트(Source)**:")
    if hidden_result.get("status") == "SUCCESS":
        hidden_count = len(hidden_result.get("hidden_sheets", []))
        very_hidden_count = len(hidden_result.get("very_hidden_sheets", []))
        total_hidden = hidden_count + very_hidden_count
        
        if total_hidden > 0:
            print(f"   - 발견된 숨김 시트: {total_hidden}개")
            if hidden_count > 0:
                print(f"     * 숨김 시트 ({hidden_count}개): {', '.join(hidden_result.get('hidden_sheets', []))}")
            if very_hidden_count > 0:
                print(f"     * 완전숨김 시트 ({very_hidden_count}개): {', '.join(hidden_result.get('very_hidden_sheets', []))}")
            print(f"     (입시 데이터는 보통 숨겨진 시트에 원본이 있습니다)")
        else:
            print(f"   - 발견된 숨김 시트: 없음")
            print(f"     (모든 시트가 보이는 상태입니다)")
    elif hidden_result.get("status") == "CRITICAL":
        print(f"   - [확인 불가] 파일이 암호화되어 있어 숨김 시트를 확인할 수 없습니다.")
        print(f"   - 오류: {hidden_result.get('error', 'N/A')}")
    else:
        print(f"   - [확인 불가] 오류 발생: {hidden_result.get('error', 'N/A')}")
    
    # 4. 최종 결론
    print(f"\n4. **최종 결론**:")
    if encryption_result.get("status") == "CRITICAL":
        print("   ⚠️  이 파일은 비밀번호 해제가 선행되어야 합니다.")
        print("   - 파일이 암호화되어 있어 Pandas로 직접 읽을 수 없습니다.")
        print("   - 비밀번호를 알고 있다면 msoffcrypto-tool을 사용하여 해제 후 분석하세요.")
    elif encryption_result.get("status") == "WARNING":
        print("   ✅ 이 파일은 Pandas로 분석 가능합니다 (암호화가 완전하지 않음).")
        print("   - 파일 암호화: 감지됨 (하지만 openpyxl/Pandas로 읽기 가능)")
        print("   - 시트 보호: 없음")
        if hidden_result.get("status") == "SUCCESS" and (len(hidden_result.get("hidden_sheets", [])) > 0 or len(hidden_result.get("very_hidden_sheets", [])) > 0):
            print("   - 숨김 시트: 발견됨 (원본 데이터가 있을 가능성 높음)")
    elif protection_result.get("status") == "SUCCESS":
        print("   ✅ 이 파일은 Pandas로 즉시 분석 가능합니다.")
        print("   - 파일 암호화: 없음")
        print("   - 시트 보호: 있음 (Pandas는 보호를 무시하고 읽을 수 있음)")
        if hidden_result.get("status") == "SUCCESS" and (len(hidden_result.get("hidden_sheets", [])) > 0 or len(hidden_result.get("very_hidden_sheets", [])) > 0):
            print("   - 숨김 시트: 발견됨 (원본 데이터가 있을 가능성 높음)")
    else:
        print("   ⚠️  파일 상태 확인 중 일부 오류가 발생했습니다.")
        print("   - 상세 오류 내용을 확인하세요.")
    
    print("\n" + "=" * 60)


def main():
    """메인 함수"""
    print("=" * 60)
    print("Excel 파일 보안 진단 도구")
    print("=" * 60)
    print(f"타겟 파일: {TARGET_FILE}")
    print(f"진단 시작 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # 파일 존재 확인
    if not os.path.exists(TARGET_FILE):
        print(f"\n[오류] 파일을 찾을 수 없습니다: {TARGET_FILE}")
        return 1
    
    # 메타데이터 수집
    metadata = get_file_metadata(TARGET_FILE)
    
    # Level 1: 파일 암호 확인
    encryption_result = check_file_encryption(TARGET_FILE)
    
    # Level 2: 시트 보호 확인
    protection_result = check_sheet_protection(TARGET_FILE)
    
    # Level 3: 숨김 시트 확인
    hidden_result = check_hidden_sheets(TARGET_FILE)
    
    # 보고서 생성
    generate_report(encryption_result, protection_result, hidden_result, metadata)
    
    print(f"\n진단 완료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return 0


if __name__ == '__main__':
    try:
        exit_code = main()
        sys.exit(exit_code)
    except KeyboardInterrupt:
        print("\n\n[중단] 사용자에 의해 중단되었습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"\n[치명적 오류] {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
